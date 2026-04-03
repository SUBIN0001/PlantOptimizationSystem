# -*- coding: utf-8 -*-
"""
excel_generator.py  –  Ashok Leyland Plant Location Decision Model
==================================================================
Generates a fully-styled 5-sheet Excel workbook with:
  • Merged title cells, freeze panes, consistent colour palette
  • Number formats: '#,##0'  '0.0000'  '0.00%'  '0.000000'
  • Conditional formatting: green→red colour scale on TOPSIS scores
                            data-bar on hybrid weights
  • Emojis: ⭐ ✅ ❌ 🔺 🔻 🔶

Sheets
------
1_Raw_Data          – raw criterion values for all locations
2_Normalised_Matrix – Min-Max normalised decision matrix
3_Weight_Calculation – AHP pairwise matrix + Entropy + Hybrid weights
4_TOPSIS_Ranking    – weighted normalised matrix, ideal solutions,
                       closeness scores, final ranking, exec summary
5_Dashboard         – bar charts (TOPSIS scores + hybrid weights) + executive summary

Usage (standalone)
------------------
    python excel_generator.py

Usage (from FastAPI)
--------------------
    from excel_generator import build_excel_bytes
    excel_bytes = build_excel_bytes(locations, pairwise_matrix)
"""

from __future__ import annotations

import io
from datetime import datetime
from typing import Dict, List

import numpy as np
import pandas as pd
from sklearn.preprocessing import MinMaxScaler

from openpyxl import Workbook
from openpyxl.styles import (Alignment, Border, Font, PatternFill, Side,
                              GradientFill)
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import (ColorScaleRule, DataBarRule,
                                       IconSetRule, CellIsRule,
                                       FormulaRule)
from openpyxl.styles.differential import DifferentialStyle

# ── Criteria (order must match backend CRITERIA_KEYS) ─────────────────────────
CRITERIA: List[Dict] = [
    {"key": "vendorBase",           "label": "Vendor Base",        "unit": "count", "isCost": False},
    {"key": "manpowerAvailability", "label": "Manpower Avail.",    "unit": "count", "isCost": False},
    {"key": "capex",                "label": "CAPEX",              "unit": "Cr",    "isCost": True},
    {"key": "govtNorms",            "label": "Govt. Norms",        "unit": "score", "isCost": False},
    {"key": "logisticsCost",        "label": "Logistics Cost",     "unit": "km",    "isCost": True},
    {"key": "economiesOfScale",     "label": "Econ. of Scale",     "unit": "score", "isCost": False},
]
KEYS   = [c["key"]   for c in CRITERIA]
LABELS = [c["label"] for c in CRITERIA]
N_CRT  = len(CRITERIA)

# ── Colour palette ─────────────────────────────────────────────────────────────
C = {
    "dark_blue":   "1F497D",
    "mid_blue":    "2F75B6",
    "light_blue":  "D9E8F5",
    "pale_blue":   "EBF3FB",
    "green":       "D9EAD3",
    "green_dark":  "6AA84F",
    "red":         "F4CCCC",
    "red_dark":    "CC0000",
    "gold":        "FFF2CC",
    "gold_dark":   "BF8F00",
    "orange":      "FCE5CD",
    "grey_bg":     "F2F2F2",
    "grey_text":   "595959",
    "white":       "FFFFFF",
    "black":       "000000",
}

# ── Borders ────────────────────────────────────────────────────────────────────
_thin_side  = Side(border_style="thin",   color="B8CCE4")
_med_side   = Side(border_style="medium", color="2F75B6")
_thick_side = Side(border_style="medium", color="1F497D")

THIN_BORDER  = Border(left=_thin_side,  right=_thin_side,  top=_thin_side,  bottom=_thin_side)
BOT_BORDER   = Border(bottom=_med_side)
OUTER_BORDER = Border(left=_thick_side, right=_thick_side,
                      top=_thick_side,  bottom=_thick_side)

# ── Style helpers ──────────────────────────────────────────────────────────────
def _font(bold=False, size=11, color=C["black"], italic=False, name="Calibri") -> Font:
    return Font(name=name, size=size, bold=bold, color=color, italic=italic)

def _fill(hex_color: str) -> PatternFill:
    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")

def _align(h="center", v="center", wrap=True) -> Alignment:
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def _set_col_widths(ws, widths: Dict[int, float]) -> None:
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w

def _apply(cell, *, font=None, fill=None, align=None, border=None, num_fmt=None):
    if font    is not None: cell.font          = font
    if fill    is not None: cell.fill          = fill
    if align   is not None: cell.alignment     = align
    if border  is not None: cell.border        = border
    if num_fmt is not None: cell.number_format = num_fmt

def _freeze(ws, row: int, col: int = 1):
    ws.freeze_panes = ws.cell(row=row, column=col)

def _row_height(ws, row: int, height: float):
    ws.row_dimensions[row].height = height


# ── Composite cell writers ─────────────────────────────────────────────────────
def _title_row(ws, text: str, row: int, ncols: int, *, size=14, bg=None):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row=row, column=1, value=text)
    _apply(c,
           font=_font(bold=True, size=size, color=C["dark_blue"]),
           fill=_fill(bg or C["pale_blue"]),
           align=_align(h="center"),
           border=BOT_BORDER)
    _row_height(ws, row, 28)


def _section_title(ws, text: str, row: int, ncols: int, *, emoji=""):
    full_text = f"{emoji}  {text}" if emoji else text
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row=row, column=1, value=full_text)
    _apply(c,
           font=_font(bold=True, size=11, color=C["mid_blue"]),
           fill=_fill(C["light_blue"]),
           align=_align(h="left"),
           border=BOT_BORDER)
    _row_height(ws, row, 20)


def _header_row(ws, headers: List[str], row: int,
                fills: List[str] | None = None) -> None:
    for ci, h in enumerate(headers, 1):
        bg = (fills[ci - 1] if fills else None) or C["dark_blue"]
        c  = ws.cell(row=row, column=ci, value=h)
        _apply(c,
               font=_font(bold=True, size=10, color=C["white"]),
               fill=_fill(bg),
               align=_align(h="center"),
               border=THIN_BORDER)
    _row_height(ws, row, 36)


def _data_cell(ws, row: int, col: int, value,
               num_fmt: str | None = None,
               align: str = "center",
               bold: bool = False,
               fill_hex: str | None = None,
               border: bool = True,
               color: str = C["black"]) -> None:
    c = ws.cell(row=row, column=col, value=value)
    _apply(c,
           font=_font(bold=bold, color=color),
           fill=_fill(fill_hex) if fill_hex else None,
           align=_align(h=align, wrap=False),
           border=THIN_BORDER if border else None,
           num_fmt=num_fmt)


def _band_row(ws, row: int, ncols: int, even: bool) -> None:
    """Alternating zebra stripe for readability."""
    bg = C["grey_bg"] if even else C["white"]
    for ci in range(1, ncols + 1):
        c = ws.cell(row=row, column=ci)
        if c.fill.patternType != "solid" or c.fill.fgColor.rgb in ("00000000", "FFFFFFFF"):
            c.fill = _fill(bg)


# ── MCDM Calculations ──────────────────────────────────────────────────────────
def _compute_mcdm(locations: List[Dict],
                  pairwise_matrix: List[List[float]],
                  alpha: float = 0.5) -> Dict:
    df = pd.DataFrame(locations)
    X  = df[KEYS].astype(float)

    # 1. Min-Max Normalisation
    scaler = MinMaxScaler()
    X_norm = pd.DataFrame(scaler.fit_transform(X), columns=KEYS)

    # 2. AHP Weights
    ahp      = np.array(pairwise_matrix, dtype=float)
    col_sums = ahp.sum(axis=0)
    norm_ahp = ahp / col_sums
    ahp_w    = norm_ahp.mean(axis=1)

    # 3. Entropy Weights
    P        = X_norm / (X_norm.sum(axis=0) + 1e-12)
    entropy  = -np.sum(P * np.log(P + 1e-12), axis=0) / np.log(max(2, len(X_norm)))
    diversity   = 1 - entropy
    entropy_w   = (diversity / diversity.sum()).values

    # 4. Hybrid Weights  (α × AHP + (1−α) × Entropy)
    hybrid_w = alpha * ahp_w + (1 - alpha) * entropy_w

    # 5. Weighted Normalised Matrix
    W = pd.DataFrame(X_norm.values * hybrid_w, columns=KEYS)

    # 6. Ideal Best / Worst
    ideal_best  = np.array([
        W.iloc[:, i].min() if CRITERIA[i]["isCost"] else W.iloc[:, i].max()
        for i in range(N_CRT)
    ])
    ideal_worst = np.array([
        W.iloc[:, i].max() if CRITERIA[i]["isCost"] else W.iloc[:, i].min()
        for i in range(N_CRT)
    ])

    d_best  = np.sqrt(((W.values - ideal_best)  ** 2).sum(axis=1))
    d_worst = np.sqrt(((W.values - ideal_worst) ** 2).sum(axis=1))
    scores  = d_worst / (d_best + d_worst + 1e-12)

    # Add original index before sorting
    df["_orig_idx"] = df.index
    df["_score"]    = scores
    ranked = df.sort_values("_score", ascending=False).reset_index(drop=True)
    ranked["_rank"] = range(1, len(ranked) + 1)

    return dict(
        raw_df=df, X_norm=X_norm,
        ahp_weights=ahp_w, entropy_weights=entropy_w, hybrid_weights=hybrid_w,
        ahp_matrix=ahp, col_sums=col_sums, norm_ahp=norm_ahp,
        W=W, ideal_best=ideal_best, ideal_worst=ideal_worst,
        d_best=d_best, d_worst=d_worst, scores=scores, ranked=ranked,
    )


# ══════════════════════════════════════════════════════════════════════════════
#  SHEET 1 — Raw Data
# ══════════════════════════════════════════════════════════════════════════════
def _create_raw_data_sheet(wb: Workbook, locations: List[Dict], alpha: float = 0.5) -> None:
    ws = wb.create_sheet("1_Raw_Data")
    NC = 7   # number of columns (Location + 6 criteria)

    _title_row(ws, "ASHOK LEYLAND  ·  PLANT LOCATION DECISION MODEL  ·  RAW DATA INPUT",
               row=1, ncols=NC)

    # Row 2 – benefit / cost direction indicators
    ind_texts = ["", "BENEFIT", "BENEFIT", "COST", "BENEFIT", "COST", "BENEFIT"]
    ind_icons = ["",  "🔺",      "🔺",      "🔻",   "🔺",      "🔻",    "🔺"]
    ind_fills = [C["dark_blue"], C["green"], C["green"],
                 C["red"],       C["green"], C["red"], C["green"]]

    for ci, (icon, txt, bg) in enumerate(zip(ind_icons, ind_texts, ind_fills), 1):
        c = ws.cell(row=2, column=ci, value=f"{icon} {txt}".strip())
        fg = C["white"] if ci == 1 else C["dark_blue"]
        _apply(c,
               font=_font(bold=True, size=9, color=fg),
               fill=_fill(bg),
               align=_align(h="center", wrap=False),
               border=THIN_BORDER)
    _row_height(ws, 2, 16)

    # Row 3 – column headers
    headers = [
        "Location",
        "Vendor Base\n(count)",
        "Manpower\nAvail. (count)",
        "CAPEX\n(Cr)",
        "Govt. Norms\n(score)",
        "Logistics\nCost (km)",
        "Economies\nof Scale",
    ]
    _header_row(ws, headers, row=3)

    # Data rows 4..
    fmts = [None, "#,##0", "#,##0", "#,##0.00", "0.0", "#,##0", "0.0"]
    for ri, loc in enumerate(locations, 4):
        vals = [
            loc.get("name", ""),
            loc.get("vendorBase", 0),
            loc.get("manpowerAvailability", 0),
            loc.get("capex", 0),
            loc.get("govtNorms", 0),
            loc.get("logisticsCost", 0),
            loc.get("economiesOfScale", 0),
        ]
        is_even = (ri % 2 == 0)
        for ci, (v, fmt) in enumerate(zip(vals, fmts), 1):
            _data_cell(ws, ri, ci, v, num_fmt=fmt,
                       align="left" if ci == 1 else "center",
                       fill_hex=C["grey_bg"] if is_even else None)
        _row_height(ws, ri, 16)

    # Legend block
    leg_r = len(locations) + 6
    _section_title(ws, "LEGEND / INSTRUCTIONS", leg_r, NC)

    legends = [
        ("🔺 BENEFIT", "Higher value is better — e.g. more suppliers means stronger local ecosystem"),
        ("🔻 COST",    "Lower value is better — e.g. lower CAPEX reduces financial risk"),
        ("Normalisation", "Min-Max normalisation applied in Sheet 2. "
                          "BENEFIT: higher norm = better. "
                          "COST: lower raw = lower norm; TOPSIS uses the min weighted value as Ideal Best."),
        ("Alpha (α)",  f"AHP influence in hybrid weight: α = {alpha}  "
                       f"({int(alpha*100)}% AHP + {int((1-alpha)*100)}% Entropy)"),
    ]
    for i, (tag, desc) in enumerate(legends):
        r = leg_r + 1 + i
        ws.cell(r, 1, tag).font      = _font(bold=True, size=10)
        ws.cell(r, 1).alignment      = _align(h="left")
        ws.cell(r, 1).border         = THIN_BORDER
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=NC)
        desc_cell = ws.cell(r, 2, desc)
        desc_cell.font      = _font(size=10, color=C["grey_text"], italic=True)
        desc_cell.alignment = _align(h="left", wrap=False)

    _set_col_widths(ws, {1: 18, 2: 16, 3: 17, 4: 13, 5: 14, 6: 15, 7: 18})
    _freeze(ws, row=4)


# ══════════════════════════════════════════════════════════════════════════════
#  SHEET 2 — Normalised Matrix
# ══════════════════════════════════════════════════════════════════════════════
def _create_normalised_sheet(wb: Workbook, data: Dict) -> None:
    ws    = wb.create_sheet("2_Normalised_Matrix")
    X_n   = data["X_norm"]
    raw   = data["raw_df"]
    n_loc = len(raw)
    NC    = N_CRT + 1

    _title_row(ws, "STEP 2  ·  MIN-MAX NORMALISED DECISION MATRIX", row=1, ncols=NC)

    formula_note = ("Formula:  norm = (x − min) / (max − min)  "
                    "|  BENEFIT: higher normalised score = better  "
                    "|  COST (CAPEX, Logistics): lower raw → lower normalised; TOPSIS uses min(weighted) as Ideal Best")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=NC)
    fn = ws.cell(row=2, column=1, value=formula_note)
    _apply(fn, font=_font(size=9, italic=True, color=C["grey_text"]),
           fill=_fill(C["pale_blue"]),
           align=_align(h="left", wrap=True))
    _row_height(ws, 2, 28)

    # Row 3 – direction labels
    dirs  = ["",  "→ norm",  "→ norm",  "→ norm (cost↓)", "→ norm", "→ norm (cost↓)", "→ norm"]
    fills = [C["dark_blue"], C["green"], C["green"],
             C["red"],       C["green"], C["red"],         C["green"]]
    for ci, (txt, bg) in enumerate(zip(dirs, fills), 1):
        c = ws.cell(row=3, column=ci, value=txt)
        fg = C["white"] if ci == 1 else C["dark_blue"]
        _apply(c, font=_font(bold=True, size=9, color=fg),
               fill=_fill(bg), align=_align(wrap=False), border=THIN_BORDER)
    _row_height(ws, 3, 16)

    # Row 4 – column headers
    _header_row(ws, ["Location"] + LABELS, row=4)

    # Data rows 5..
    for ri in range(n_loc):
        r    = ri + 5
        name = raw.iloc[ri].get("name", f"loc_{ri+1}")
        _data_cell(ws, r, 1, name, align="left",
                   fill_hex=C["grey_bg"] if ri % 2 else None)
        for ci, key in enumerate(KEYS, 2):
            v = float(X_n.iloc[ri][key])
            _data_cell(ws, r, ci, round(v, 6), num_fmt="0.000000",
                       fill_hex=C["grey_bg"] if ri % 2 else None)
        _row_height(ws, r, 15)

    # Conditional formatting: green=high, red=low on each criterion column
    data_start = 5
    data_end   = 4 + n_loc
    for ci in range(2, NC + 1):
        col_letter = get_column_letter(ci)
        rng = f"{col_letter}{data_start}:{col_letter}{data_end}"
        ws.conditional_formatting.add(rng, ColorScaleRule(
            start_type="min",  start_color="F4CCCC",   # red  = low
            mid_type="num",    mid_value=0.5, mid_color="FFEB84",  # yellow = mid
            end_type="max",    end_color="63BE7B",      # green = high
        ))

    # Summary statistics
    sum_r = data_end + 2
    _section_title(ws, "Summary Statistics (normalised values)", sum_r, NC)
    for i, (label, func) in enumerate([("MIN",  X_n.min()),
                                        ("MAX",  X_n.max()),
                                        ("MEAN", X_n.mean()),
                                        ("STD",  X_n.std())]):
        r = sum_r + 1 + i
        ws.cell(r, 1, label).font      = _font(bold=True, size=10)
        ws.cell(r, 1).alignment        = _align(h="right")
        ws.cell(r, 1).border           = THIN_BORDER
        ws.cell(r, 1).fill             = _fill(C["light_blue"])
        for ci, key in enumerate(KEYS, 2):
            _data_cell(ws, r, ci, round(float(func[key]), 6), num_fmt="0.000000",
                       fill_hex=C["light_blue"])

    _set_col_widths(ws, {1: 18, **{i: 15 for i in range(2, NC + 1)}})
    _freeze(ws, row=5)


# ══════════════════════════════════════════════════════════════════════════════
#  SHEET 3 — Weight Calculation
# ══════════════════════════════════════════════════════════════════════════════
def _create_weights_sheet(wb: Workbook, data: Dict, alpha: float) -> None:
    ws   = wb.create_sheet("3_Weight_Calculation")
    ahp  = data["ahp_matrix"]
    cs   = data["col_sums"]
    na   = data["norm_ahp"]
    aw   = data["ahp_weights"]
    ew   = data["entropy_weights"]
    hw   = data["hybrid_weights"]

    # ── Section A: AHP Pairwise Matrix ────────────────────────────────────────
    NC_A = N_CRT + 1
    _title_row(ws,
               f"STEP 3  ·  WEIGHT CALCULATION  —  AHP · ENTROPY · HYBRID  (α = {alpha})",
               row=1, ncols=NC_A)
    _section_title(ws, "A.  AHP PAIRWISE COMPARISON MATRIX", row=2, ncols=NC_A, emoji="📐")

    _header_row(ws, ["Criterion  ↓ / →"] + LABELS, row=3)

    for ri, label in enumerate(LABELS):
        r = ri + 4
        ws.cell(r, 1, label).font      = _font(bold=True, size=10)
        ws.cell(r, 1).alignment        = _align(h="left")
        ws.cell(r, 1).border           = THIN_BORDER
        ws.cell(r, 1).fill             = _fill(C["pale_blue"])
        for ci in range(N_CRT):
            v    = float(ahp[ri, ci])
            diag = (ri == ci)
            _data_cell(ws, r, ci + 2, round(v, 4), num_fmt="0.0000",
                       fill_hex=C["light_blue"] if diag else (C["grey_bg"] if ri % 2 else None))
        _row_height(ws, r, 15)

    # Column sums row
    cs_r = 4 + N_CRT
    ws.cell(cs_r, 1, "Column Sums:").font      = _font(bold=True, size=10, color=C["mid_blue"])
    ws.cell(cs_r, 1).border                    = THIN_BORDER
    ws.cell(cs_r, 1).fill                      = _fill(C["light_blue"])
    ws.cell(cs_r, 1).alignment                 = _align(h="right")
    for ci, v in enumerate(cs, 2):
        _data_cell(ws, cs_r, ci, round(float(v), 4), num_fmt="0.0000",
                   fill_hex=C["light_blue"], bold=True)

    # Normalised AHP matrix (collapsed sub-section)
    na_r = cs_r + 2
    _section_title(ws, "A2.  NORMALISED AHP MATRIX  (column-normalised; row mean = AHP weight)",
                   row=na_r, ncols=NC_A, emoji="📊")
    _header_row(ws, ["Criterion"] + LABELS, row=na_r + 1)
    for ri, label in enumerate(LABELS):
        r = na_r + 2 + ri
        ws.cell(r, 1, label).font      = _font(bold=True, size=10)
        ws.cell(r, 1).alignment        = _align(h="left")
        ws.cell(r, 1).border           = THIN_BORDER
        ws.cell(r, 1).fill             = _fill(C["pale_blue"])
        for ci in range(N_CRT):
            v = float(na[ri, ci])
            _data_cell(ws, r, ci + 2, round(v, 6), num_fmt="0.000000",
                       fill_hex=C["grey_bg"] if ri % 2 else None)
        _row_height(ws, r, 15)

    # ── Section B: Weights Summary ─────────────────────────────────────────────
    sec_b = na_r + 2 + N_CRT + 2
    _section_title(ws, "B.  WEIGHTS SUMMARY  — AHP · Entropy · Hybrid", sec_b, 9, emoji="⚖️")

    w_hdrs = ["Criterion", "Type",
              "AHP Weight", "AHP %",
              "Entropy Weight", "Entropy %",
              "Hybrid Weight", "Hybrid %",
              "Rank"]
    _header_row(ws, w_hdrs, sec_b + 1)

    sorted_idx = np.argsort(hw)[::-1]
    rank_map   = {i: rank + 1 for rank, i in enumerate(sorted_idx)}

    for i, crit in enumerate(CRITERIA):
        r         = sec_b + 2 + i
        is_cost   = crit["isCost"]
        type_str  = "🔻 COST" if is_cost else "🔺 BENEFIT"
        type_fill = C["red"] if is_cost else C["green"]
        is_top    = rank_map[i] == 1
        row_fill  = C["gold"] if is_top else (C["grey_bg"] if i % 2 else None)

        row_vals = [
            crit["label"], type_str,
            float(aw[i]),  float(aw[i]),         # AHP, AHP%
            float(ew[i]),  float(ew[i]),          # Entropy, Entropy%
            float(hw[i]),  float(hw[i]),          # Hybrid, Hybrid%
            rank_map[i],
        ]
        fmts = [None, None,
                "0.000000", "0.00%",
                "0.000000", "0.00%",
                "0.0000",   "0.00%",
                "0"]
        for ci, (v, fmt) in enumerate(zip(row_vals, fmts), 1):
            cell_fill = type_fill if ci == 2 else row_fill
            weight_hi = (ci in (7, 8)) and is_top          # highlight best hybrid
            _data_cell(ws, r, ci, v, num_fmt=fmt,
                       align="left" if ci == 1 else "center",
                       fill_hex=cell_fill if cell_fill else None,
                       bold=is_top,
                       color=C["dark_blue"] if weight_hi else C["black"])
        _row_height(ws, r, 15)

    # Totals row
    tot_r = sec_b + 2 + N_CRT
    ws.cell(tot_r, 1, "TOTAL:").font      = _font(bold=True, size=10)
    ws.cell(tot_r, 1).border             = THIN_BORDER
    ws.cell(tot_r, 1).fill              = _fill(C["light_blue"])
    ws.cell(tot_r, 1).alignment         = _align(h="right")
    for col, vals in [(3, aw), (5, ew), (7, hw)]:
        _data_cell(ws, tot_r, col, round(float(sum(vals)), 6),
                   num_fmt="0.000000", bold=True, fill_hex=C["light_blue"])
    for col in (4, 6, 8):            # percent totals = 100%
        _data_cell(ws, tot_r, col, 1.0, num_fmt="0.00%",
                   bold=True, fill_hex=C["light_blue"])

    # Data-bar conditional formatting on Hybrid Weight column
    hw_col_letter = get_column_letter(7)
    hw_range = f"{hw_col_letter}{sec_b+2}:{hw_col_letter}{tot_r-1}"
    ws.conditional_formatting.add(hw_range, DataBarRule(
        start_type="min", end_type="max",
        color="2F75B6",
    ))

    _set_col_widths(ws, {1: 22, 2: 12, 3: 13, 4: 10,
                         5: 14, 6: 10, 7: 13, 8: 10, 9: 7})
    _set_col_widths(ws, {i: 13 for i in range(10, N_CRT + 2)})
    _freeze(ws, row=4)


# ══════════════════════════════════════════════════════════════════════════════
#  SHEET 4 — TOPSIS Ranking
# ══════════════════════════════════════════════════════════════════════════════
def _create_topsis_sheet(wb: Workbook, data: Dict, alpha: float = 0.5) -> None:
    ws      = wb.create_sheet("4_TOPSIS_Ranking")
    W       = data["W"]
    ib      = data["ideal_best"]
    iw_arr  = data["ideal_worst"]
    db      = data["d_best"]
    dw      = data["d_worst"]
    scores  = data["scores"]
    ranked  = data["ranked"]
    n_loc   = len(ranked)
    NC      = N_CRT + 1

    _title_row(ws, "STEP 4  ·  TOPSIS RANKING  —  FINAL SCORES & OPTIMAL PLANT LOCATION",
               row=1, ncols=NC, bg=C["pale_blue"])

    # ── Section A: Weighted Normalised Matrix ──────────────────────────────────
    _section_title(ws, "A.  WEIGHTED NORMALISED MATRIX   V = w × normalised",
                   row=2, ncols=NC, emoji="🔢")
    _header_row(ws, ["Location"] + LABELS, row=3)

    data_start_a = 4
    for ri, row_data in ranked.iterrows():
        r        = ri + data_start_a
        orig_idx = int(row_data.get("_orig_idx", ri))
        rank_val = int(row_data["_rank"])
        is_top   = rank_val == 1
        bg       = C["gold"] if is_top else (C["grey_bg"] if ri % 2 else None)

        _data_cell(ws, r, 1, row_data["name"], align="left", bold=is_top, fill_hex=bg)
        for ci, key in enumerate(KEYS, 2):
            v = float(W.iloc[orig_idx, ci - 2]) if orig_idx < len(W) else 0.0
            _data_cell(ws, r, ci, round(v, 6), num_fmt="0.000000",
                       fill_hex=bg)
        _row_height(ws, r, 15)
    data_end_a = data_start_a + n_loc - 1

    # ── Section B: Ideal Solutions ─────────────────────────────────────────────
    sec_b = data_end_a + 2
    _section_title(ws, "B.  IDEAL SOLUTIONS   A* = ideal best   |   A⁻ = ideal worst",
                   sec_b, NC, emoji="⚡")
    _header_row(ws, [""] + LABELS, sec_b + 1)

    for (label, vals, bg) in [
        ("✅  A*  Ideal Best",  ib,     C["green"]),
        ("❌  A⁻  Ideal Worst", iw_arr, C["red"]),
    ]:
        r = sec_b + 2 + (0 if "Best" in label else 1)
        ws.cell(r, 1, label).font      = _font(bold=True, size=10)
        ws.cell(r, 1).fill             = _fill(bg)
        ws.cell(r, 1).border           = THIN_BORDER
        ws.cell(r, 1).alignment        = _align(h="left")
        for ci, v in enumerate(vals, 2):
            _data_cell(ws, r, ci, round(float(v), 6), num_fmt="0.000000",
                       fill_hex=bg)
        _row_height(ws, r, 15)

    # ── Section C: Closeness Scores & Final Ranking ────────────────────────────
    sec_c = sec_b + 5
    _section_title(ws, "C.  TOPSIS CLOSENESS SCORES & FINAL RANKING",
                   sec_c, 7, emoji="🏆")

    score_hdrs = ["Rank", "Location",
                  "D⁺\n(dist ideal best)",
                  "D⁻\n(dist ideal worst)",
                  "C*\n(Closeness Score)",
                  "Status",
                  "Recommendation"]
    _header_row(ws, score_hdrs, sec_c + 1)

    data_start_c = sec_c + 2
    for ri, row_data in ranked.iterrows():
        r        = data_start_c + ri
        rank_val = int(row_data["_rank"])
        orig_idx = int(row_data.get("_orig_idx", ri))
        score    = float(scores[orig_idx]) if orig_idx < len(scores) else 0.0
        d_b      = float(db[orig_idx])     if orig_idx < len(db)     else 0.0
        d_w      = float(dw[orig_idx])     if orig_idx < len(dw)     else 0.0

        is_top = rank_val == 1
        bg     = C["gold"] if is_top else (C["grey_bg"] if ri % 2 else None)
        bg_c   = C["gold"] if is_top else (C["green"] if rank_val <= 3 else C["orange"])

        if rank_val == 1:
            rec = "⭐ OPTIMAL — Recommended"
            status = "⭐ #1"
        elif rank_val <= 3:
            rec = "✅ Strong Candidate"
            status = f"✅ #{rank_val}"
        else:
            rec = "🔶 Viable Option"
            status = f"🔶 #{rank_val}"

        row_vals = [rank_val, row_data["name"], d_b, d_w, score, status, rec]
        fmts     = ["0", None, "0.000000", "0.000000", "0.000000", None, None]
        aligns   = ["center", "left", "center", "center", "center", "center", "left"]
        for ci, (v, fmt, aln) in enumerate(zip(row_vals, fmts, aligns), 1):
            cell_bg = bg_c if ci == 6 else bg
            _data_cell(ws, r, ci, v, num_fmt=fmt, align=aln,
                       bold=is_top, fill_hex=cell_bg)
        _row_height(ws, r, 16)
    data_end_c = data_start_c + n_loc - 1

    # Conditional formatting – colour scale on C* column (col 5)
    c_star_col = get_column_letter(5)
    ws.conditional_formatting.add(
        f"{c_star_col}{data_start_c}:{c_star_col}{data_end_c}",
        ColorScaleRule(
            start_type="min",  start_color="F4CCCC",
            mid_type="percentile", mid_value=50, mid_color="FFEB84",
            end_type="max",    end_color="63BE7B",
        ),
    )

    # ── Section D: Executive Summary ──────────────────────────────────────────
    sec_d     = data_end_c + 3
    best_row  = ranked.iloc[0]
    best_name = best_row["name"]
    best_orig = int(best_row.get("_orig_idx", 0))
    best_sc   = float(scores[best_orig])

    _section_title(ws, "D.  EXECUTIVE SUMMARY", sec_d, 7, emoji="📋")

    summary_items = [
        ("⭐  Optimal Location",    best_name),
        ("C* Score",                f"{best_sc:.6f}"),
        ("2nd Best Location",       ranked.iloc[1]["name"] if n_loc > 1 else "—"),
        ("3rd Best Location",       ranked.iloc[2]["name"] if n_loc > 2 else "—"),
        ("Locations Evaluated",     str(n_loc)),
        ("Criteria Used",           f"{N_CRT} (Hybrid AHP+Entropy weights, α=0.5)"),
        ("Method",                  "TOPSIS — Technique for Order of Preference by Similarity to Ideal Solution"),
        ("Generated",               datetime.now().strftime("%Y-%m-%d  %H:%M:%S")),
    ]
    for i, (label, value) in enumerate(summary_items):
        r = sec_d + 1 + i
        is_opt  = i == 0
        row_bg  = C["gold"]       if is_opt else C["pale_blue"]
        val_clr = C["dark_blue"]  if is_opt else C["black"]
        val_sz  = 12              if is_opt else 10

        ws.cell(r, 1, label).font      = _font(bold=True, size=10)
        ws.cell(r, 1).alignment        = _align(h="left")
        ws.cell(r, 1).border           = THIN_BORDER
        ws.cell(r, 1).fill             = _fill(row_bg)

        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=7)
        c = ws.cell(r, 2, value)
        _apply(c,
               font=_font(bold=is_opt, size=val_sz, color=val_clr),
               fill=_fill(row_bg),
               align=_align(h="left"),
               border=THIN_BORDER)
        _row_height(ws, r, 18 if is_opt else 15)

    # Recommendation paragraph
    rec_row = sec_d + len(summary_items) + 2
    rank_2  = ranked.iloc[1]["name"] if n_loc > 1 else "—"
    ws.merge_cells(start_row=rec_row, start_column=1, end_row=rec_row, end_column=7)
    rec_text = (
        f"RECOMMENDATION:  Following comprehensive TOPSIS analysis with hybrid "
        f"AHP+Entropy weights (α = {alpha}), {best_name} achieves the highest "
        f"closeness coefficient of {best_sc:.6f}, making it the optimal plant "
        f"location for Ashok Leyland's expansion. {rank_2} is the recommended "
        f"alternative if site-specific constraints arise. Both CAPEX and Logistics "
        f"Cost (cost criteria) were appropriately penalised, while Vendor Base, "
        f"Manpower Availability, Govt. Norms and Economies of Scale (benefit "
        f"criteria) were maximised in the composite score."
    )
    c = ws.cell(rec_row, 1, rec_text)
    _apply(c,
           font=_font(size=10, color=C["dark_blue"], italic=True),
           fill=_fill(C["light_blue"]),
           align=Alignment(horizontal="left", vertical="top", wrap_text=True),
           border=OUTER_BORDER)
    _row_height(ws, rec_row, 68)

    _set_col_widths(ws, {1: 10, 2: 22, 3: 18, 4: 18, 5: 18, 6: 14, 7: 30})
    _freeze(ws, row=4)


# ══════════════════════════════════════════════════════════════════════════════
#  SHEET 5 — Dashboard  (bar chart + executive summary)
# ══════════════════════════════════════════════════════════════════════════════
def _create_dashboard_sheet(wb: Workbook, data: Dict, locations: List[Dict], alpha: float) -> None:
    from openpyxl.chart import BarChart, Reference
    from openpyxl.chart.series import DataPoint
    from openpyxl.drawing.fill import PatternFillProperties

    ws = wb.create_sheet("5_Dashboard")
    ws.sheet_view.showGridLines = False   # cleaner look

    ranked  = data["ranked"]
    scores  = data["scores"]
    hw      = data["hybrid_weights"]
    n_loc   = len(ranked)

    # Colour gradient for bars: rank 1 = gold, 2 = green, 3 = teal, rest = blue shades
    BAR_COLORS = [
        "FFD700",  # rank 1  – gold
        "4CAF50",  # rank 2  – green
        "00BCD4",  # rank 3  – cyan/teal
        "2196F3",  # rank 4  – blue
        "9C27B0",  # rank 5  – purple
        "FF5722",  # rank 6  – deep orange
        "795548",  # rank 7  – brown
        "607D8B",  # rank 8  – blue-grey
        "F44336",  # rank 9  – red
        "009688",  # rank 10 – teal
    ]

    # ── TITLE BAND (rows 1-3) ──────────────────────────────────────────────────
    ws.merge_cells("A1:O1")
    title_cell = ws["A1"]
    title_cell.value    = "ASHOK LEYLAND  ·  PLANT LOCATION DECISION  ·  EXECUTIVE DASHBOARD"
    title_cell.font     = Font(name="Calibri", size=18, bold=True, color="FFFFFF")
    title_cell.fill     = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 40

    ws.merge_cells("A2:O2")
    sub_cell = ws["A2"]
    sub_cell.value     = (f"TOPSIS + Hybrid AHP-Entropy  |  α = {alpha}  "
                          f"|  {n_loc} Locations  |  {N_CRT} Criteria  "
                          f"|  Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    sub_cell.font      = Font(name="Calibri", size=10, italic=True, color="D9E8F5")
    sub_cell.fill      = PatternFill(start_color="2F75B6", end_color="2F75B6", fill_type="solid")
    sub_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 18
    ws.row_dimensions[3].height = 8   # spacer

    # ── KPI CARDS (row 4-8, columns A-O split into 4 cards) ───────────────────
    # Card 1: Optimal Location
    best      = ranked.iloc[0]
    best_orig = int(best.get("_orig_idx", 0))
    best_sc   = float(scores[best_orig])
    rank2     = ranked.iloc[1]["name"] if n_loc > 1 else "—"
    rank3     = ranked.iloc[2]["name"] if n_loc > 2 else "—"
    feasible_count = n_loc   # all ranked are feasible

    cards = [
        # (col_start, col_end, bg_color, label, value, sub)
        ("A", "C",  "FFF2CC", "OPTIMAL LOCATION",    "FFF8E1",
         best["name"],        f"C* = {best_sc:.4f}",  "BF8F00", "7F6000"),
        ("E", "G",  "D9EAD3", "2nd BEST",             "F0FFF0",
         rank2,               "Strong Candidate",     "6AA84F", "2D6A4F"),
        ("I", "K",  "D9E8F5", "3rd BEST",             "EBF3FB",
         rank3,               "Strong Candidate",     "2F75B6", "1F497D"),
        ("M", "O",  "F4CCCC", "LOCATIONS EVALUATED",  "FFF0F0",
         str(n_loc),          f"{N_CRT} criteria used", "CC0000", "990000"),
    ]

    for (c1, c2, header_bg, label, body_bg, value, sub, hdr_fg, body_fg) in cards:
        # header row
        ws.merge_cells(f"{c1}4:{c2}4")
        hc = ws[f"{c1}4"]
        hc.value     = label
        hc.font      = Font(name="Calibri", size=9, bold=True, color=hdr_fg)
        hc.fill      = PatternFill(start_color=header_bg, end_color=header_bg, fill_type="solid")
        hc.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[4].height = 16

        # value row
        ws.merge_cells(f"{c1}5:{c2}7")
        vc = ws[f"{c1}5"]
        vc.value     = value
        vc.font      = Font(name="Calibri", size=16, bold=True, color=body_fg)
        vc.fill      = PatternFill(start_color=body_bg, end_color=body_bg, fill_type="solid")
        vc.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for rr in (5, 6, 7):
            ws.row_dimensions[rr].height = 16

        # sub-label row
        ws.merge_cells(f"{c1}8:{c2}8")
        sc2 = ws[f"{c1}8"]
        sc2.value     = sub
        sc2.font      = Font(name="Calibri", size=9, italic=True, color=hdr_fg)
        sc2.fill      = PatternFill(start_color=header_bg, end_color=header_bg, fill_type="solid")
        sc2.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[8].height = 14

    ws.row_dimensions[9].height = 10  # spacer

    # ── DATA TABLE for charts (rows 11.., hidden columns A:C) ─────────────────
    # TOPSIS Scores table: col A=Location, col B=C* Score, col C=Rank
    ws.cell(10, 1, "Location").font  = _font(bold=True)
    ws.cell(10, 2, "C* Score").font  = _font(bold=True)
    ws.cell(10, 3, "Rank").font      = _font(bold=True)
    ws.cell(10, 4, "Criterion").font = _font(bold=True)
    ws.cell(10, 5, "Hybrid W").font  = _font(bold=True)

    for ri, row_data in ranked.iterrows():
        r        = ri + 11
        orig_idx = int(row_data.get("_orig_idx", ri))
        sc       = float(scores[orig_idx])
        ws.cell(r, 1, row_data["name"])
        ws.cell(r, 2, round(sc, 6))
        ws.cell(r, 2).number_format = "0.000000"
        ws.cell(r, 3, int(row_data["_rank"]))

    for ci, (crit, w) in enumerate(zip(CRITERIA, hw)):
        r = ci + 11
        ws.cell(r, 4, crit["label"])
        ws.cell(r, 5, round(float(w), 6))
        ws.cell(r, 5).number_format = "0.000000"

    # Hide the data rows (they exist only for chart references)
    # openpyxl doesn't support row hide cleanly; just narrow the rows
    for r in range(10, 11 + max(n_loc, N_CRT)):
        ws.row_dimensions[r].height = 0.1

    # ── BAR CHART 1: TOPSIS C* Scores ─────────────────────────────────────────
    chart1 = BarChart()
    chart1.type       = "col"          # vertical columns
    chart1.grouping   = "clustered"
    chart1.varyColors = True           # each bar gets its own colour
    chart1.title      = "TOPSIS Closeness Scores (C*)  —  Location Ranking"
    chart1.style      = 2
    chart1.y_axis.title  = "Closeness Score (C*)"
    chart1.x_axis.title  = "Plant Location"
    chart1.y_axis.scaling.min = 0
    chart1.y_axis.numFmt  = "0.000"
    chart1.dataLabels         = None
    chart1.width  = 22   # cm
    chart1.height = 14   # cm

    # Data series: C* scores (col B, rows 11..10+n_loc)
    data_ref = Reference(ws, min_col=2, min_row=10,
                         max_col=2, max_row=10 + n_loc)
    cats_ref = Reference(ws, min_col=1, min_row=11,
                         max_col=1, max_row=10 + n_loc)
    chart1.add_data(data_ref, titles_from_data=True)
    chart1.set_categories(cats_ref)

    # Colour individual bars by rank  (openpyxl DataPoint)
    series = chart1.series[0]
    series.graphicalProperties.solidFill = BAR_COLORS[0]   # default = gold
    for rank_i in range(n_loc):
        # ranked is sorted rank 1..n; rank_i=0 → rank 1
        hex_col = BAR_COLORS[rank_i] if rank_i < len(BAR_COLORS) else "607D8B"
        pt = DataPoint(idx=rank_i)
        pt.graphicalProperties.solidFill = hex_col
        series.dPt.append(pt)

    # Place chart starting at row 11, col E (column 5)
    chart1.anchor = "E10"
    ws.add_chart(chart1)

    # ── BAR CHART 2: Hybrid Weights ────────────────────────────────────────────
    chart2 = BarChart()
    chart2.type       = "bar"          # horizontal bars
    chart2.grouping   = "clustered"
    chart2.varyColors = True
    chart2.title      = "Hybrid Criterion Weights  (AHP + Entropy, α = " + str(alpha) + ")"
    chart2.style      = 3
    chart2.x_axis.title = "Weight"
    chart2.y_axis.title = "Criterion"
    chart2.x_axis.numFmt = "0.0%"
    chart2.width  = 15
    chart2.height = 12

    w_data = Reference(ws, min_col=5, min_row=10,
                       max_col=5, max_row=10 + N_CRT)
    w_cats = Reference(ws, min_col=4, min_row=11,
                       max_col=4, max_row=10 + N_CRT)
    chart2.add_data(w_data, titles_from_data=True)
    chart2.set_categories(w_cats)

    # Colour bars for weight chart
    WEIGHT_COLORS = ["2196F3", "4CAF50", "F44336", "FF9800", "9C27B0", "00BCD4"]
    series2 = chart2.series[0]
    series2.graphicalProperties.solidFill = WEIGHT_COLORS[0]
    for i in range(N_CRT):
        hex_col = WEIGHT_COLORS[i % len(WEIGHT_COLORS)]
        pt2 = DataPoint(idx=i)
        pt2.graphicalProperties.solidFill = hex_col
        series2.dPt.append(pt2)

    chart2.anchor = "A30"
    ws.add_chart(chart2)

    # ── EXECUTIVE SUMMARY BLOCK (to the right of weight chart) ────────────────
    # Place at approx row 30, col I
    summ_col   = 9    # column I
    summ_start = 30

    def _kpi_row(ws, row, col, label, value, bg, fg="000000", val_size=13):
        ws.merge_cells(
            start_row=row, start_column=col,
            end_row=row,   end_column=col + 1
        )
        lc = ws.cell(row, col, label)
        lc.font      = Font(name="Calibri", size=10, bold=True, color=fg)
        lc.fill      = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
        lc.alignment = Alignment(horizontal="left", vertical="center")
        lc.border    = THIN_BORDER

        ws.merge_cells(
            start_row=row, start_column=col + 2,
            end_row=row,   end_column=col + 5
        )
        vc = ws.cell(row, col + 2, value)
        vc.font      = Font(name="Calibri", size=val_size, bold=True, color=fg)
        vc.fill      = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
        vc.alignment = Alignment(horizontal="center", vertical="center")
        vc.border    = THIN_BORDER
        ws.row_dimensions[row].height = 20

    # Title
    ws.merge_cells(
        start_row=summ_start, start_column=summ_col,
        end_row=summ_start,   end_column=summ_col + 5
    )
    hdr = ws.cell(summ_start, summ_col, "EXECUTIVE SUMMARY")
    hdr.font      = Font(name="Calibri", size=13, bold=True, color="FFFFFF")
    hdr.fill      = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
    hdr.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[summ_start].height = 28

    kpi_rows = [
        ("Optimal Location",    f"⭐  {best['name']}",          "FFF2CC", "7F6000", 14),
        ("TOPSIS Score (C*)",   f"{best_sc:.6f}",               "D9EAD3", "2D6A4F", 13),
        ("2nd Best",            f"✅  {rank2}",                  "D9E8F5", "1F497D", 12),
        ("3rd Best",            f"✅  {rank3}",                  "EBF3FB", "2F75B6", 12),
        ("Locations Evaluated", str(n_loc),                     "F2F2F2", "000000", 12),
        ("Criteria Used",       f"{N_CRT}  (AHP+Entropy)",      "F2F2F2", "000000", 12),
        ("AHP Weight (α)",      f"{int(alpha*100)}%",           "FCE5CD", "7F3300", 12),
        ("Entropy Weight (β)",  f"{int((1-alpha)*100)}%",       "FCE5CD", "7F3300", 12),
        ("Generated",           datetime.now().strftime("%Y-%m-%d  %H:%M"), "F2F2F2", "595959", 10),
    ]
    for i, (label, value, bg, fg, vsz) in enumerate(kpi_rows):
        _kpi_row(ws, summ_start + 1 + i, summ_col, label, value, bg, fg, vsz)

    # Recommendation paragraph
    rec_r = summ_start + len(kpi_rows) + 2
    ws.merge_cells(
        start_row=rec_r, start_column=summ_col,
        end_row=rec_r + 3, end_column=summ_col + 5
    )
    rec = ws.cell(rec_r, summ_col,
        f"RECOMMENDATION:  Based on TOPSIS with hybrid AHP+Entropy "
        f"weights (α={alpha}), {best['name']} achieves the highest closeness "
        f"coefficient of {best_sc:.6f} and is recommended as the optimal "
        f"plant location. {rank2} is the preferred alternative site."
    )
    rec.font      = Font(name="Calibri", size=10, italic=True, color=C["dark_blue"])
    rec.fill      = PatternFill(start_color="EBF3FB", end_color="EBF3FB", fill_type="solid")
    rec.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    rec.border    = OUTER_BORDER
    for rr in range(rec_r, rec_r + 4):
        ws.row_dimensions[rr].height = 16

    # Column widths
    col_widths = {1: 18, 2: 12, 3: 12, 4: 3,
                  5: 20, 6: 18, 7: 18, 8: 18, 9: 20,
                  10: 3, 11: 20, 12: 14, 13: 14, 14: 14, 15: 14}
    _set_col_widths(ws, col_widths)


# ══════════════════════════════════════════════════════════════════════════════
#  CONSTRAINT FILTERING HELPER
# ══════════════════════════════════════════════════════════════════════════════
def _apply_constraints(locations: List[Dict],
                       constraints: List[Dict],
                       region_filter: Dict | None = None) -> tuple:
    """
    Split locations into feasible / infeasible based on active constraints.

    Returns (feasible_locs, infeasible_locs, active_constraints)
    where infeasible_locs is a list of (loc_dict, [failed_constraint_labels]).
    """
    active = [c for c in constraints if c.get("enabled", False)]

    # Apply region / state filter first
    rf_enabled  = region_filter.get("regionFilterEnabled", False) if region_filter else False
    sel_regions = region_filter.get("selectedRegions", [])        if region_filter else []

    feasible   = []
    infeasible = []   # list of (loc, [reason_strings])

    for loc in locations:
        reasons = []

        # Region / State filter
        # A location passes if its region OR its state is in sel_regions.
        # We check both independently so state-only selections work correctly
        # even when a location also has a region value.
        if rf_enabled and sel_regions:
            loc_region = loc.get("region") or ""
            loc_state  = loc.get("state")  or ""
            if loc_region not in sel_regions and loc_state not in sel_regions:
                which = f"region='{loc_region}'" if loc_region else f"state='{loc_state}'"
                reasons.append(f"{which} not in selected filter: {sel_regions}")

        # Numeric constraints
        for c in active:
            val = float(loc.get(c["key"], 0))
            op  = c["operator"]
            thr = float(c["value"])
            if op == "gte" and val < thr:
                reasons.append(f"{c['label']} {val:.2f} < {thr:.2f} (need ≥)")
            elif op == "lte" and val > thr:
                reasons.append(f"{c['label']} {val:.2f} > {thr:.2f} (need ≤)")
            elif op == "eq" and abs(val - thr) > 0.001:
                reasons.append(f"{c['label']} {val:.2f} ≠ {thr:.2f} (need =)")

        if reasons:
            infeasible.append((loc, reasons))
        else:
            feasible.append(loc)

    return feasible, infeasible, active


# ══════════════════════════════════════════════════════════════════════════════
#  SHEET 0 — Constraints Summary
# ══════════════════════════════════════════════════════════════════════════════
def _create_constraints_sheet(wb: Workbook,
                              all_locations: List[Dict],
                              feasible: List[Dict],
                              infeasible: List[tuple],
                              active_constraints: List[Dict],
                              region_filter: Dict | None = None) -> None:
    ws = wb.create_sheet("0_Constraints", 0)   # insert as first sheet
    NC = 5

    # ── Title ─────────────────────────────────────────────────────────────────
    _title_row(ws, "STEP 0  ·  FEASIBILITY CONSTRAINTS  —  FILTER APPLIED BEFORE MCDM",
               row=1, ncols=NC)

    # ── Summary stats row ─────────────────────────────────────────────────────
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=NC)
    summary_text = (
        f"Total locations uploaded: {len(all_locations)}   |   "
        f"Feasible (passed all constraints): {len(feasible)}   |   "
        f"Infeasible (filtered out): {len(infeasible)}   |   "
        f"Active constraints: {len(active_constraints)}"
    )
    sc = ws.cell(row=2, column=1, value=summary_text)
    _apply(sc,
           font=_font(size=10, italic=True, color=C["grey_text"]),
           fill=_fill(C["pale_blue"]),
           align=_align(h="left", wrap=False))
    _row_height(ws, 2, 18)

    # ── Section A: Active Numeric Constraints ─────────────────────────────────
    _section_title(ws, "A.  ACTIVE FEASIBILITY CONSTRAINTS", row=3, ncols=NC, emoji="⚙️")

    rf_enabled  = region_filter.get("regionFilterEnabled", False) if region_filter else False
    sel_regions = region_filter.get("selectedRegions", [])        if region_filter else []

    hdrs = ["Criterion", "Operator", "Threshold", "Unit", "Impact"]
    _header_row(ws, hdrs, row=4)

    cur_row = 5
    if rf_enabled and sel_regions:
        _data_cell(ws, cur_row, 1, "Region / State", align="left", bold=True)
        _data_cell(ws, cur_row, 2, "is in")
        _data_cell(ws, cur_row, 3, ", ".join(sel_regions), align="left")
        _data_cell(ws, cur_row, 4, "—")
        _data_cell(ws, cur_row, 5, "Locations outside this list are excluded", align="left")
        _row_height(ws, cur_row, 15)
        cur_row += 1

    UNIT_MAP = {
        "vendorBase": "count", "manpowerAvailability": "count",
        "capex": "Cr", "govtNorms": "score",
        "logisticsCost": "km", "economiesOfScale": "score",
    }
    OP_LABEL  = {"gte": "≥", "lte": "≤", "eq": "="}

    if not active_constraints and not (rf_enabled and sel_regions):
        ws.merge_cells(start_row=cur_row, start_column=1,
                       end_row=cur_row, end_column=NC)
        c = ws.cell(cur_row, 1, "⚠  No constraints were enabled — all locations passed to MCDM.")
        _apply(c,
               font=_font(size=11, bold=True, color=C["gold_dark"]),
               fill=_fill(C["gold"]),
               align=_align(h="left"))
        _row_height(ws, cur_row, 20)
        cur_row += 1
    else:
        for i, c in enumerate(active_constraints):
            r       = cur_row + i
            unit    = UNIT_MAP.get(c["key"], "")
            op_lbl  = OP_LABEL.get(c["operator"], c["operator"])
            is_cost = c.get("isCost", False)
            bg      = C["grey_bg"] if i % 2 else None
            _data_cell(ws, r, 1, c.get("label", c["key"]), align="left", bold=True,
                       fill_hex=bg)
            _data_cell(ws, r, 2, op_lbl, fill_hex=bg)
            _data_cell(ws, r, 3, float(c["value"]),
                       num_fmt="#,##0.00", fill_hex=bg)
            _data_cell(ws, r, 4, unit, fill_hex=bg)
            direction = "COST — lower is better" if is_cost else "BENEFIT — higher is better"
            _data_cell(ws, r, 5, direction, align="left", fill_hex=bg)
            _row_height(ws, r, 15)
        cur_row += len(active_constraints)

    cur_row += 2   # spacer

    # ── Section B: Infeasible Locations ───────────────────────────────────────
    _section_title(ws, "B.  INFEASIBLE LOCATIONS  (excluded from MCDM ranking)",
                   row=cur_row, ncols=NC, emoji="❌")
    cur_row += 1

    _header_row(ws, ["#", "Location", "Region", "State", "Reason(s) for Exclusion"],
                row=cur_row)
    cur_row += 1

    if not infeasible:
        ws.merge_cells(start_row=cur_row, start_column=1,
                       end_row=cur_row, end_column=NC)
        c = ws.cell(cur_row, 1, "✅  All locations passed the feasibility constraints.")
        _apply(c,
               font=_font(size=10, italic=True, color=C["green_dark"]),
               fill=_fill(C["green"]),
               align=_align(h="left"))
        _row_height(ws, cur_row, 16)
        cur_row += 1
    else:
        for i, (loc, reasons) in enumerate(infeasible):
            r = cur_row + i
            _data_cell(ws, r, 1, i + 1, fill_hex=C["red"])
            _data_cell(ws, r, 2, loc.get("name", ""), align="left",
                       bold=True, fill_hex=C["red"])
            _data_cell(ws, r, 3, loc.get("region") or "—",
                       align="left", fill_hex=C["red"])
            _data_cell(ws, r, 4, loc.get("state") or "—",
                       align="left", fill_hex=C["red"])
            _data_cell(ws, r, 5, "  |  ".join(reasons),
                       align="left", fill_hex=C["red"])
            _row_height(ws, r, 15)
        cur_row += len(infeasible) + 2

    # ── Section C: Feasible Locations ─────────────────────────────────────────
    _section_title(ws, "C.  FEASIBLE LOCATIONS  (included in MCDM ranking)",
                   row=cur_row, ncols=NC, emoji="✅")
    cur_row += 1

    _header_row(ws,
                ["#", "Location", "Region", "State",
                 *[c.get("label", c["key"]) for c in active_constraints[:3]]]
                if active_constraints
                else ["#", "Location", "Region", "State", "All Constraints Passed"],
                row=cur_row)
    cur_row += 1

    for i, loc in enumerate(feasible):
        r  = cur_row + i
        bg = C["grey_bg"] if i % 2 else None
        _data_cell(ws, r, 1, i + 1, fill_hex=bg)
        _data_cell(ws, r, 2, loc.get("name", ""), align="left", bold=True, fill_hex=bg)
        _data_cell(ws, r, 3, loc.get("region") or "—", align="left", fill_hex=bg)
        _data_cell(ws, r, 4, loc.get("state") or "—", align="left", fill_hex=bg)

        if active_constraints:
            for ci, c in enumerate(active_constraints[:3]):
                val = float(loc.get(c["key"], 0))
                _data_cell(ws, r, 5 + ci, val, num_fmt="#,##0.00",
                           fill_hex=C["green"])
        else:
            _data_cell(ws, r, 5, "✅ Passed",
                       fill_hex=C["green"])
        _row_height(ws, r, 15)

    _set_col_widths(ws, {1: 5, 2: 22, 3: 18, 4: 18, 5: 50})
    _freeze(ws, row=5)


# ══════════════════════════════════════════════════════════════════════════════
#  PUBLIC API
# ══════════════════════════════════════════════════════════════════════════════
def build_excel_bytes(locations: List[Dict],
                      pairwise_matrix: List[List[float]],
                      alpha: float = 0.5,
                      constraints: List[Dict] | None = None,
                      region_filter: Dict | None = None) -> bytes:
    """
    Run full MCDM pipeline and return a styled 6-sheet .xlsx as bytes.

    Parameters
    ----------
    locations       : list of location dicts (same schema as FastAPI Location model)
    pairwise_matrix : 6×6 AHP pairwise comparison matrix
    alpha           : AHP weight in hybrid (default 0.5 = 50% AHP, 50% Entropy)
    constraints     : list of constraint dicts with keys:
                      key, label, operator ('gte'/'lte'/'eq'), value, enabled
    region_filter   : dict with regionFilterEnabled (bool) and selectedRegions (list)

    Returns
    -------
    bytes — ready to be set as FastAPI Response content
    """
    if constraints is None:
        constraints = []

    # ── 1. Apply feasibility constraints ───────────────────────────────────────
    feasible, infeasible, active_constraints = _apply_constraints(
        locations, constraints, region_filter
    )

    # Need at least one feasible location for MCDM
    locs_for_mcdm = feasible if feasible else locations
    data = _compute_mcdm(locs_for_mcdm, pairwise_matrix, alpha=alpha)

    # ── 2. Build workbook ──────────────────────────────────────────────────────
    wb = Workbook()
    wb.remove(wb.active)    # remove blank default sheet

    # Sheet 0: Constraints (prepended)
    _create_constraints_sheet(wb, locations, feasible, infeasible,
                              active_constraints, region_filter)
    # Sheet 1‑5: MCDM results (uses feasible locations only)
    _create_raw_data_sheet(wb, locs_for_mcdm, alpha)
    _create_normalised_sheet(wb, data)
    _create_weights_sheet(wb, data, alpha)
    _create_topsis_sheet(wb, data, alpha)
    _create_dashboard_sheet(wb, data, locs_for_mcdm, alpha)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ══════════════════════════════════════════════════════════════════════════════
#  STANDALONE DEMO
# ══════════════════════════════════════════════════════════════════════════════
if __name__ == "__main__":
    sample_locations = [
        {"id": "1", "name": "Chennai",    "vendorBase": 350, "manpowerAvailability": 85,  "capex": 2.8, "govtNorms": 4.0, "logisticsCost": 20,  "economiesOfScale": 4.0},
        {"id": "2", "name": "Pune",       "vendorBase": 450, "manpowerAvailability": 90,  "capex": 3.2, "govtNorms": 3.5, "logisticsCost": 45,  "economiesOfScale": 4.5},
        {"id": "3", "name": "Ahmedabad",  "vendorBase": 280, "manpowerAvailability": 70,  "capex": 2.5, "govtNorms": 4.2, "logisticsCost": 55,  "economiesOfScale": 3.8},
        {"id": "4", "name": "Hosur",      "vendorBase": 320, "manpowerAvailability": 80,  "capex": 2.2, "govtNorms": 3.8, "logisticsCost": 35,  "economiesOfScale": 3.5},
        {"id": "5", "name": "Nagpur",     "vendorBase": 200, "manpowerAvailability": 65,  "capex": 1.8, "govtNorms": 3.2, "logisticsCost": 80,  "economiesOfScale": 3.0},
        {"id": "6", "name": "Coimbatore", "vendorBase": 300, "manpowerAvailability": 75,  "capex": 2.4, "govtNorms": 3.6, "logisticsCost": 40,  "economiesOfScale": 3.8},
        {"id": "7", "name": "Hyderabad",  "vendorBase": 380, "manpowerAvailability": 88,  "capex": 3.0, "govtNorms": 4.5, "logisticsCost": 50,  "economiesOfScale": 4.2},
        {"id": "8", "name": "Bangalore",  "vendorBase": 420, "manpowerAvailability": 95,  "capex": 3.8, "govtNorms": 4.0, "logisticsCost": 60,  "economiesOfScale": 4.8},
    ]
    sample_ahp = [
        [1.0,  2.0,  3.0,  3.0,  2.0,  4.0],
        [0.5,  1.0,  2.0,  2.0,  2.0,  3.0],
        [0.333,0.5,  1.0,  1.0,  0.5,  2.0],
        [0.333,0.5,  1.0,  1.0,  0.5,  2.0],
        [0.5,  0.5,  2.0,  2.0,  1.0,  2.0],
        [0.25, 0.333,0.5,  0.5,  0.5,  1.0],
    ]

    out = build_excel_bytes(sample_locations, sample_ahp, alpha=0.5)
    fname = f"Ashok_Leyland_Results_{datetime.now().strftime('%Y-%m-%d')}.xlsx"
    with open(fname, "wb") as fh:
        fh.write(out)
    print(f"[OK] Saved: {fname}  ({len(out):,} bytes)")
